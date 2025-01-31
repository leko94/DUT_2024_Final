import dash
from dash import dcc, html
from dash.dependencies import Input, Output
import pandas as pd
import plotly.express as px
import plotly.graph_objs as go
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from sklearn.impute import SimpleImputer

# Initialize the Dash app
app = dash.Dash(__name__, suppress_callback_exceptions=True)
server = app.server  # Expose the server for WSGI

# File paths to the Excel files
staff_file_path = 'Chart in Microsoft PowerPoint.xlsx'
students_file_path = 'Students.xlsx'
student_performance_file_path = '/content/Student Perfomances.xlsx'

# Functions to create charts for staff data
def create_staff_charts():
    workbook = load_workbook(staff_file_path, data_only=True)
    sheet_names = workbook.sheetnames

    # Extract data from each sheet
    titles = {}
    x_labels = {}
    y_labels = {}
    dfs = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        titles[sheet_name] = sheet['A1'].value
        df = pd.read_excel(staff_file_path, sheet_name=sheet_name, header=None, skiprows=0)
        x_labels[sheet_name] = df.iloc[1, 0]  # X-axis label from A2
        y_labels[sheet_name] = df.iloc[1, 1]  # Y-axis label from B2
        dfs[sheet_name] = df.iloc[2:, [0, 1]].rename(columns={0: x_labels[sheet_name], 1: y_labels[sheet_name]})

    # Create bar charts
    def create_bar_chart(df, title, x_label, y_label, color):
        fig = px.bar(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        fig.update_layout(
            autosize=False,
            width=600,
            height=600,
            bargap=0.2
        )
        fig.update_traces(text=df[df.columns[1]], textposition='outside')  # Add values on bars
        return fig

    # Create line charts
    def create_line_chart(df, title, x_label, y_label, color):
        fig = px.line(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        fig.update_layout(
            autosize=False,
            width=600,
            height=600
        )
        return fig

    # Create forecast charts
    def create_forecast_chart(df, title, x_label, y_label, color):
        df[x_label] = pd.to_numeric(df[x_label], errors='coerce')
        df[y_label] = pd.to_numeric(df[y_label], errors='coerce')
        
        # Linear Regression for Forecasting
        X = df[[x_label]].values.reshape(-1, 1)
        y = df[y_label].values
        model = LinearRegression().fit(X, y)
        
        # Forecast for 2030
        future_years = np.array([2030]).reshape(-1, 1)
        forecast_values = model.predict(future_years)
        
        fig = px.line(
            df,
            x=df.columns[0],
            y=df.columns[1],
            title=title,
            labels={df.columns[0]: x_label, df.columns[1]: y_label},
            color_discrete_sequence=[color]
        )
        forecast_df = pd.DataFrame({x_label: [2030], y_label: forecast_values})
        fig.add_scatter(x=forecast_df[x_label], y=forecast_df[y_label], mode='markers+text', text=['2030'], textposition='top center', marker=dict(color='red', size=10))
        fig.update_layout(
            autosize=False,
            width=600,
            height=600
        )
        return fig

    # Define custom colors for each chart
    colors = ['blue', 'green', 'orange', 'purple']

    # Create figures for each sheet
    bar_figures = {}
    line_figures = {}
    forecast_figures = {}

    for i, sheet_name in enumerate(sheet_names):
        color = colors[i % len(colors)]
        df = dfs[sheet_name]
        title = titles[sheet_name]
        x_label = x_labels[sheet_name]
        y_label = y_labels[sheet_name]
        
        if sheet_name in ['Sheet1', 'Sheet2', 'Sheet3']:
            bar_figures[sheet_name] = create_bar_chart(df, title, x_label, y_label, color)
            forecast_figures[sheet_name] = create_forecast_chart(df, f"Forecast for 2030 - {title}", x_label, y_label, color)
        elif sheet_name == 'Sheet4':
            line_figures[sheet_name] = create_line_chart(df, title, x_label, y_label, color)

    # Load data for Sheet5
    df_sheet5 = pd.read_excel(staff_file_path, sheet_name='Sheet5')

    # Extract year columns and data for Sheet5
    years = df_sheet5.columns[1:10]
    departments = df_sheet5.iloc[:, 0]
    data = df_sheet5.iloc[:, 1:10]

    # Convert data to appropriate format for Sheet5
    df_data = data.copy()
    df_data.columns = years
    df_data['Department'] = departments.values
    df_data = df_data.melt(id_vars='Department', var_name='Year', value_name='Percentage')

    # Extract 2014 and 2022 data
    df_2014 = df_data[df_data['Year'] == '2014'][['Department', 'Percentage']].set_index('Department')
    df_2022 = df_data[df_data['Year'] == '2022'][['Department', 'Percentage']].set_index('Department')

    # Calculate the difference between 2022 and 2014
    percentage_diff = df_2022.join(df_2014, lsuffix='_2022', rsuffix='_2014')
    percentage_diff['Percentage Difference'] = percentage_diff['Percentage_2022'] - percentage_diff['Percentage_2014']

    # Reset index and prepare data for the graph
    percentage_diff = percentage_diff.reset_index()
    percentage_diff.columns = ['Department', 'Percentage 2022', 'Percentage 2014', 'Percentage Difference']

    # Create the bar graph showing the percentage difference
    fig_diff = px.bar(
        percentage_diff,
        x='Department',
        y='Percentage Difference',
        title='Percentage Difference (2022 vs. 2014) for Academic Staff with PhD',
        labels={'Percentage Difference': 'Percentage Difference'},
        height=600,
        color='Department'
    )
    fig_diff.update_traces(text=percentage_diff['Percentage Difference'], textposition='outside')
    fig_diff.update_layout(
        legend_title='Departments',
        autosize=False,
        width=800,
        height=600
    )

    # Create the bar graph for Sheet5 without text labels
    fig_sheet5 = px.bar(
        df_data,
        x='Year',
        y='Percentage',
        color='Department',
        title='Percentage of Full-Time Permanent Academic Staff with PhD (2014-2022)',
        labels={'Percentage': 'Percentage'},
        height=600
    )
    fig_sheet5.update_traces(textposition='none')  # Remove text annotations

    return bar_figures, line_figures, forecast_figures, fig_diff, fig_sheet5

# Functions to create charts for students data
def create_students_charts():
    # Load data from Sheet1
    df1 = pd.read_excel(students_file_path, sheet_name='Sheet1')
    df1.columns = ['Year', 'Actual', 'Planned']
    df1['Actual'] = pd.to_numeric(df1['Actual'], errors='coerce')
    df1['Planned'] = pd.to_numeric(df1['Planned'], errors='coerce')
    df1 = df1.dropna()

    # Create the first line chart
    fig1 = px.line(
        df1,
        x='Year',
        y=['Actual', 'Planned'],
        labels={'value': 'Headcount', 'variable': 'Type'},
        title='Headcount Enrolment: Planned vs Achieved (2014-2022)',
        markers=True
    )
    fig1.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Create the second chart with linear regression forecast
    def create_linear_regression_forecast_chart(df):
        df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
        df['Actual'] = pd.to_numeric(df['Actual'], errors='coerce')
        df = df.dropna()
        X = df[['Year']].values.reshape(-1, 1)
        y = df['Actual'].values
        model = LinearRegression().fit(X, y)
        future_years = np.array([2030]).reshape(-1, 1)
        forecast_value = model.predict(future_years)[0]
        fig2 = px.line(
            df,
            x='Year',
            y='Actual',
            title='Linear Regression Forecast for 2030',
            labels={'Year': 'Year', 'Actual': 'Headcount'},
            markers=True
        )
        fig2.add_scatter(
            x=[2030],
            y=[forecast_value],
            mode='markers+text',
            text=['2030'],
            textposition='top right',
            marker=dict(color='red', size=10),
            name='Forecast'
        )
        return fig2

    fig2 = create_linear_regression_forecast_chart(df1)

    # Load data from Sheet2 and get the title from cell A1
    df2 = pd.read_excel(students_file_path, sheet_name='Sheet2', header=1)  # Skip the header row
    title2 = pd.read_excel(students_file_path, sheet_name='Sheet2', header=None).iloc[0, 0]  # Get the title from cell A1

    # Compute the Difference
    df2['Difference'] = df2['Actual'] - df2['Planned']

    # Create the bar graph for Sheet2
    fig3 = px.bar(
        df2,
        x='Departments',
        y=['Planned', 'Actual'],
        title=title2,
        labels={'value': 'Number of Students', 'variable': 'Type'},
        text='Difference'  # Display the difference on the bars
    )
    fig3.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Load data from Sheet3
    df3 = pd.read_excel(students_file_path, sheet_name='Sheet3', header=None)

    # Extract title and relevant columns
    title3 = df3.iloc[0, 0]  # Title from cell A1
    df3.columns = ['Department', '2014', '2022', 'Ignore', 'Growth']  # Set column names manually
    df3 = df3.iloc[2:]  # Skip rows before actual data
    df3 = df3[['Department', '2014', '2022', 'Growth']]  # Select relevant columns

    # Convert columns to numeric values and round the Growth values
    df3['2014'] = pd.to_numeric(df3['2014'], errors='coerce')
    df3['2022'] = pd.to_numeric(df3['2022'], errors='coerce')
    df3['Growth'] = pd.to_numeric(df3['Growth'], errors='coerce').round(2)

    # Melt the DataFrame for better plotting
    df3_melted = df3.melt(id_vars='Department', value_vars=['2014', '2022'], var_name='Year', value_name='Value')

    # Create the bar graph for Sheet3
    fig4 = px.bar(
        df3_melted,
        x='Department',
        y='Value',
        color='Year',
        text=df3_melted['Department'].map(df3.set_index('Department')['Growth']),  # Add rounded Growth as text
        title=title3,
        labels={'Value': 'Number of Students', 'Year': 'Year'},
        color_discrete_map={'2014': 'blue', '2022': 'green'}
    )
    fig4.update_traces(texttemplate='%{text:.2f}', textposition='outside')
    fig4.update_layout(
        autosize=False,
        width=800,
        height=600,
        legend_title_text='Year'
    )

    # Load data from Sheet4
    df4 = pd.read_excel(students_file_path, sheet_name='Sheet4')

    # Extract title and labels
    title4 = '% African Students'  # Title from cell A1
    x_label4 = 'Year'  # X-axis label from cell A2
    y_label4 = '% African Students'  # Y-axis label from cell B2

    # Set headers and extract data
    df4.columns = [x_label4, y_label4]
    df4 = df4.iloc[1:]  # Skip rows before actual data

    # Convert '% African Students' to float after stripping '%'
    df4[y_label4] = df4[y_label4].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet4
    fig5 = px.line(
        df4,
        x=x_label4,
        y=y_label4,
        title=title4,
        labels={x_label4: 'Year', y_label4: '% African Students'},
        markers=True
    )
    fig5.update_layout(
        autosize=False,
        width=800,
        height=600
    )

    # Load data from Sheet5
    df5 = pd.read_excel(students_file_path, sheet_name='Sheet5')

    # Extract title and labels
    title5 = "Percentage Female Students"  # Title from cell A1
    x_label5 = 'Year'  # X-axis label from cell A2
    y_label5 = 'Percentage Female Students'  # Y-axis label from cell B2

    # Set headers and extract data
    df5.columns = [x_label5, y_label5]
    df5 = df5.iloc[1:]  # Skip rows before actual data

    # Convert '% Female Students' to float after stripping '%'
    df5[y_label5] = df5[y_label5].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet5 with a different color
    fig6 = px.line(
        df5,
        x=x_label5,
        y=y_label5,
        title=title5,
        labels={x_label5: 'Year', y_label5: 'Percentage Female Students'},
        markers=True,
        line_shape='linear'
    )
    fig6.update_layout(
        autosize=False,
        width=800,
        height=600,
        plot_bgcolor='lightgray'  # Different color for distinction
    )

    # Load data from Sheet6
    df6 = pd.read_excel(students_file_path, sheet_name='Sheet6')

    # Extract title and labels
    title6 = "Faculty Postgraduate Enrolment"  # Title from cell A1
    x_label6 = 'Year'  # X-axis label from cell A2
    y_label6 = '% Enrolment'  # Y-axis label from cell B2

    # Set headers and extract data
    df6.columns = [x_label6, y_label6]
    df6 = df6.iloc[1:]  # Skip rows before actual data

    # Convert '% Enrolment' to float after stripping '%'
    df6[y_label6] = df6[y_label6].astype(str).str.rstrip('%').astype(float)

    # Create the line graph
    fig7 = go.Figure()

    # Add the line trace
    fig7.add_trace(go.Scatter(
        x=df6[x_label6],
        y=df6[y_label6],
        mode='lines',
        line=dict(color='purple'),  # Set the color to purple
        name='Enrolment Percentage'
    ))

    # Update layout for the graph from Sheet6
    fig7.update_layout(
        title=title6,
        xaxis_title=x_label6,
        yaxis_title=y_label6,
        autosize=False,
        width=800,
        height=600,
        plot_bgcolor='lightgray'  # Different color for distinction
    )

    # Load data from Sheet7
    df7 = pd.read_excel(students_file_path, sheet_name='Sheet7')
    df7.columns = df7.columns.str.strip()

    # Create the bar chart with actual percentages on top of each bar
    fig8 = px.bar(
        df7, 
        x='Department', 
        y=['UG', 'PG upto Masters', 'PG'], 
        title='Enrolment by Level',
        labels={'value': 'Percentage', 'variable': 'Enrolment Level'},
        barmode='group',
        text_auto=True  # Add text_auto=True to display values on bars
    )

    # Load data from Sheet8
    df8 = pd.read_excel(students_file_path, sheet_name='Sheet8')
    df8.columns = df8.columns.str.strip()

    # Melt the DataFrame for better plotting
    df8_melted = df8.melt(id_vars='Departments', value_vars=df8.columns[1:-1], var_name='Year', value_name='Percentage')
    df8_melted['Percentage'] = df8_melted['Percentage'].astype(str).str.rstrip('%').astype(float)

    # Create the line graph for Sheet8
    fig9 = px.line(
        df8_melted,
        x='Year',
        y='Percentage',
        color='Departments',
        title='Postgraduate (M+D) Enrolment',
        markers=True
    )
    fig9.update_layout(
        autosize=False,
        width=800,
        height=600,
        legend_title_text='Departments'
    )

    # Create the bar graph for Departments and Difference 2014 vs 2022
    fig10 = px.bar(
        df8,
        x='Departments',
        y='Difference 2014 vs 2022',
        title='Difference 2014 vs 2022 by Department',
        labels={'Departments': 'Departments', 'Difference 2014 vs 2022': 'Difference'},
        text='Difference 2014 vs 2022'
    )
    fig10.update_layout(
        autosize=False,
        width=800,
        height=600
    )
    fig10.update_traces(texttemplate='%{text:.2f}', textposition='outside')

    # Load data from Sheet9
    df9 = pd.read_excel(students_file_path, sheet_name='Sheet9')
    df9.columns = df9.columns.str.strip()

    # Melt the DataFrame for better plotting
    df9_melted = df9.melt(id_vars='Derpatnment', value_vars=df9.columns[1:], var_name='Year', value_name='No. of Postgraduate enrolment')

    # Create the bar graph for Sheet9
    fig11 = px.bar(
        df9_melted,
        x='Year',
        y='No. of Postgraduate enrolment',
        color='Derpatnment',
        title='Postgraduate Enrolment - Actual Student Numbers',
        text='No. of Postgraduate enrolment',
        barmode='group'
    )
    fig11.update_layout(
        autosize=False,
        width=1000,  # Increased the width
        height=700,  # Increased the height
        legend_title_text='Departments'
    )
    fig11.update_traces(texttemplate='%{text}', textposition='outside')

    # Load data from Sheet10
    df10 = pd.read_excel(students_file_path, sheet_name='Sheet10')
    df10.columns = df10.columns.str.strip()

    # Melt the DataFrame for better plotting
    df10_melted = df10.melt(id_vars='Department', value_vars=df10.columns[1:], var_name='Year', value_name='Percentage')
    df10_melted['Percentage'] = df10_melted['Percentage'].astype(str).str.rstrip('%').astype(float)

    # Create the bar graph for Sheet10
    fig12 = px.bar(
        df10_melted,
        x='Year',
        y='Percentage',
        color='Department',
        title='International student Postgraduate enrolment',
        text='Percentage',
        barmode='group'
    )
    fig12.update_layout(
        autosize=False,
        width=1200,  # Increased the width
        height=800,  # Increased the height
        legend_title_text='Departments'
    )
    fig12.update_traces(texttemplate='%{text}', textposition='outside')

    # Load data from Sheet11
    df11 = pd.read_excel(students_file_path, sheet_name='Sheet11')
    df11.columns = df11.columns.str.strip()

    # Melt the DataFrame for better plotting
    df11_melted = df11.melt(id_vars='Department', value_vars=df11.columns[1:], var_name='Year', value_name='No. of Postgraduate enrolment')

    # Create the bar graph for Sheet11
    fig13 = px.bar(
        df11_melted,
        x='Year',
        y='No. of Postgraduate enrolment',
        color='Department',
        title='International Students Postgraduate Enrolment - Actual Numbers',
        text='No. of Postgraduate enrolment',
        barmode='group'
    )
    fig13.update_layout(
        autosize=False,
        width=1200,  # Increased the width
        height=800,  # Increased the height
        legend_title_text='Departments'
    )
    fig13.update_traces(texttemplate='%{text}', textposition='outside')

    return fig1, fig2, fig3, fig4, fig5, fig6, fig7, fig8, fig9, fig10, fig11, fig12, fig13

# Functions to create charts for student performance data
def create_student_performance_charts():
    # Read the Excel files into DataFrames
    df1 = pd.read_excel(student_performance_file_path, sheet_name='Sheet1')
    df2 = pd.read_excel(student_performance_file_path, sheet_name='Sheet2')
    df3 = pd.read_excel(student_performance_file_path, sheet_name='Sheet3')
    df4 = pd.read_excel(student_performance_file_path, sheet_name='Sheet4')
    df5 = pd.read_excel(student_performance_file_path, sheet_name='Sheet5')
    df6 = pd.read_excel(student_performance_file_path, sheet_name='Sheet6')
    df7 = pd.read_excel(student_performance_file_path, sheet_name='Sheet7')
    df8 = pd.read_excel(student_performance_file_path, sheet_name='Sheet8')
    df9 = pd.read_excel(student_performance_file_path, sheet_name='Sheet9')
    df10 = pd.read_excel(student_performance_file_path, sheet_name='Sheet10')
    df11 = pd.read_excel(student_performance_file_path, sheet_name='Sheet11')
    df12 = pd.read_excel(student_performance_file_path, sheet_name='Sheet12')
    df13 = pd.read_excel(student_performance_file_path, sheet_name='Sheet13')
    df14 = pd.read_excel(student_performance_file_path, sheet_name='Sheet14')
    df15 = pd.read_excel(student_performance_file_path, sheet_name='Sheet15')

    # Ensure Success Rates in Sheet1 are strings and convert them to float
    df1['Success Rates'] = df1['Success Rates'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet1
    X1 = df1[['Year']]
    y1 = df1['Success Rates']

    # Impute any missing values in X1 and y1
    imputer_X1 = SimpleImputer(strategy='mean')
    imputer_y1 = SimpleImputer(strategy='mean')
    X1 = imputer_X1.fit_transform(X1)
    y1 = imputer_y1.fit_transform(y1.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet1
    model1 = LinearRegression()
    model1.fit(X1, y1)

    # Forecasting for the next 5 years for Sheet1
    future_years1 = np.arange(df1['Year'].max() + 1, df1['Year'].max() + 6).reshape(-1, 1)
    predictions1 = model1.predict(future_years1)

    # Append predictions to the DataFrame for Sheet1
    future_df1 = pd.DataFrame({'Year': future_years1.flatten(), 'Success Rates': predictions1})

    # Ensure Success Rates in Sheet2 are strings and convert them to float
    for year in ['2019', '2020', '2021', '2022']:
        df2[year] = df2[year].astype(str).str.rstrip('%').astype(float)

    # Melt the DataFrame from Sheet2 to have long-form data for easier plotting
    df_melted2 = df2.melt(id_vars=['Department'], value_vars=['2019', '2020', '2021', '2022'],
                        var_name='Year', value_name='Success Rates')

    # Ensure Success Rates in Sheet3 are strings and convert them to float
    df3['Success Rates of First Time Entering Students'] = df3['Success Rates of First Time Entering Students'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet3
    X3 = df3[['Year']]
    y3 = df3['Success Rates of First Time Entering Students']

    # Impute any missing values in X3 and y3
    imputer_X3 = SimpleImputer(strategy='mean')
    imputer_y3 = SimpleImputer(strategy='mean')
    X3 = imputer_X3.fit_transform(X3)
    y3 = imputer_y3.fit_transform(y3.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet3
    model3 = LinearRegression()
    model3.fit(X3, y3)

    # Forecasting for the next 5 years for Sheet3
    future_years3 = np.arange(df3['Year'].max() + 1, df3['Year'].max() + 6).reshape(-1, 1)
    predictions3 = model3.predict(future_years3)

    # Append predictions to the DataFrame for Sheet3
    future_df3 = pd.DataFrame({'Year': future_years3.flatten(), 'Success Rates': predictions3})

    # Ensure Success Rates in Sheet4 are strings and convert them to float
    df4['Success Rates of African Students'] = df4['Success Rates of African Students'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet4
    X4 = df4[['Year']]
    y4 = df4['Success Rates of African Students']

    # Impute any missing values in X4 and y4
    imputer_X4 = SimpleImputer(strategy='mean')
    imputer_y4 = SimpleImputer(strategy='mean')
    X4 = imputer_X4.fit_transform(X4)
    y4 = imputer_y4.fit_transform(y4.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet4
    model4 = LinearRegression()
    model4.fit(X4, y4)

    # Forecasting for the next 5 years for Sheet4
    future_years4 = np.arange(df4['Year'].max() + 1, df4['Year'].max() + 6).reshape(-1, 1)
    predictions4 = model4.predict(future_years4)

    # Append predictions to the DataFrame for Sheet4
    future_df4 = pd.DataFrame({'Year': future_years4.flatten(), 'Success Rates': predictions4})

    # Ensure Success Rates in Sheet5 are strings and convert them to float
    df5['FACULTY'] = df5['FACULTY'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet6 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022']:
        df6[year] = df6[year].astype(str).str.rstrip('%').astype(float)
    df6['Difference: 2014 vs 2022'] = df6['Difference: 2014 vs 2022'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet7 are strings and convert them to float
    for year in ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022']:
        df7[year] = df7[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet8 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022']:
        df8[year] = df8[year].astype(str).str.rstrip('%').astype(float)
    df8['Difference: 2014 vs 2022'] = df8['Difference: 2014 vs 2022'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet9 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020']:
        df9[year] = df9[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet10 are strings and convert them to float
    df10['Dropout'] = df10['Dropout'].astype(str).str.rstrip('%').astype(float)
    df10['Throughput'] = df10['Throughput'].astype(str).str.rstrip('%').astype(float)
    df10['Still in Progress'] = df10['Still in Progress'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet11 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2019', '2020', '2021', '2022']:
        df11[year] = df11[year].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet12 are strings and convert them to float
    df12['Faculty'] = df12['Faculty'].astype(str).str.rstrip('%').astype(float)

    # Filter out the row containing "Difference: 2014 vs. 2021"
    df12_filtered = df12[df12['Year'] != 'Difference: 2014 vs. 2021']
    df12_filtered['Year'] = df12_filtered['Year'].astype(int)

    # Prepare data for Linear Regression from Sheet12
    X12 = df12_filtered[['Year']]
    y12 = df12_filtered['Faculty']

    # Impute any missing values in X12 and y12
    imputer_X12 = SimpleImputer(strategy='mean')
    imputer_y12 = SimpleImputer(strategy='mean')
    X12 = imputer_X12.fit_transform(X12)
    y12 = imputer_y12.fit_transform(y12.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet12
    model12 = LinearRegression()
    model12.fit(X12, y12)

    # Forecasting for the next 5 years for Sheet12
    future_years12 = np.arange(df12_filtered['Year'].max() + 1, df12_filtered['Year'].max() + 6).reshape(-1, 1)
    predictions12 = model12.predict(future_years12)

    # Append predictions to the DataFrame for Sheet12
    future_df12 = pd.DataFrame({'Year': future_years12.flatten(), 'Faculty': predictions12})

    # Ensure Success Rates in Sheet13 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022']:
        df13[year] = df13[year].astype(str).str.rstrip('%').astype(float)
    df13['Difference: 2014 vs 2022'] = df13['Difference: 2014 vs 2022'].astype(str).str.rstrip('%').astype(float)

    # Ensure Success Rates in Sheet14 are strings and convert them to float
    df14['Faculty'] = df14['Faculty'].astype(str).str.rstrip('%').astype(float)

    # Prepare data for Linear Regression from Sheet14
    X14 = df14[['Year']]
    y14 = df14['Faculty']

    # Impute any missing values in X14 and y14
    imputer_X14 = SimpleImputer(strategy='mean')
    imputer_y14 = SimpleImputer(strategy='mean')
    X14 = imputer_X14.fit_transform(X14)
    y14 = imputer_y14.fit_transform(y14.values.reshape(-1, 1)).ravel()

    # Fit the Linear Regression model for Sheet14
    model14 = LinearRegression()
    model14.fit(X14, y14)

    # Forecasting for the next 5 years for Sheet14
    future_years14 = np.arange(df14['Year'].max() + 1, df14['Year'].max() + 6).reshape(-1, 1)
    predictions14 = model14.predict(future_years14)

    # Append predictions to the DataFrame for Sheet14
    future_df14 = pd.DataFrame({'Year': future_years14.flatten(), 'Faculty': predictions14})

    # Ensure Success Rates in Sheet15 are strings and convert them to float
    for year in ['2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022']:
        df15[year] = df15[year].astype(str).str.rstrip('%').astype(float)

    # Find the year with the highest overall performance in Sheet2
    highest_year = df_melted2.loc[df_melted2['Success Rates'].idxmax()]['Year']

    # Create the graphs for each sheet
    graphs = [
        dcc.Graph(
            id='success-rate-graph-sheet1',
            figure={
                'data': [
                    go.Scatter(
                        x=df1['Year'],
                        y=df1['Success Rates'],
                        mode='lines+markers',
                        name='Actual Success Rate',
                        line=dict(color='blue')
                    ),
                    go.Scatter(
                        x=future_df1['Year'],
                        y=future_df1['Success Rates'],
                        mode='lines+markers',
                        name='Forecasted Success Rate',
                        line=dict(color='red', dash='dash')
                    )
                ],
                'layout': go.Layout(
                    title='FAS Overall Student Success Rate (Sheet1)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'}
                )
            }
        ),
        dcc.Graph(
            id='success-rate-bar-graph-sheet2',
            figure={
                'data': [
                    go.Bar(
                        x=df_melted2[df_melted2['Department'] == dept]['Year'],
                        y=df_melted2[df_melted2['Department'] == dept]['Success Rates'],
                        name=dept
                    ) for dept in df2['Department']
                ],
                'layout': go.Layout(
                    title='Department Success Rates by Year (Sheet2)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='highlight-year',
            figure={
                'data': [
                    go.Bar(
                        x=df_melted2[df_melted2['Department'] == dept]['Year'],
                        y=df_melted2[df_melted2['Department'] == dept]['Success Rates'],
                        name=dept
                    ) for dept in df2['Department']
                ],
                'layout': go.Layout(
                    title='Department Success Rates by Year with Highlight (Sheet2)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'},
                    barmode='group',
                    annotations=[
                        dict(
                            x=highest_year,
                            y=df_melted2['Success Rates'].max(),
                            xref='x',
                            yref='y',
                            text='Highest Performance Year',
                            showarrow=True,
                            arrowhead=7,
                            ax=0,
                            ay=-40
                        )
                    ]
                )
            }
        ),
        dcc.Graph(
            id='success-rate-graph-sheet3',
            figure={
                'data': [
                    go.Scatter(
                        x=df3['Year'],
                        y=df3['Success Rates of First Time Entering Students'],
                        mode='lines+markers',
                        name='Actual Success Rate',
                        line=dict(color='green')
                    ),
                    go.Scatter(
                        x=future_df3['Year'],
                        y=future_df3['Success Rates'],
                        mode='lines+markers',
                        name='Forecasted Success Rate',
                        line=dict(color='orange', dash='dash')
                    )
                ],
                'layout': go.Layout(
                    title='Success Rates of First Time Entering Students (Sheet3)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'}
                )
            }
        ),
        dcc.Graph(
            id='success-rate-graph-sheet4',
            figure={
                'data': [
                    go.Scatter(
                        x=df4['Year'],
                        y=df4['Success Rates of African Students'],
                        mode='lines+markers',
                        name='Actual Success Rate',
                        line=dict(color='purple')
                    ),
                    go.Scatter(
                        x=future_df4['Year'],
                        y=future_df4['Success Rates'],
                        mode='lines+markers',
                        name='Forecasted Success Rate',
                        line=dict(color='brown', dash='dash')
                    )
                ],
                'layout': go.Layout(
                    title='Success Rates of African Students (Sheet4)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'}
                )
            }
        ),
        dcc.Graph(
            id='success-rate-graph-sheet5',
            figure={
                'data': [
                    go.Scatter(
                        x=df5['Year'],
                        y=df5['FACULTY'],
                        mode='lines+markers',
                        name='FACULTY Success Rate',
                        line=dict(color='cyan')
                    )
                ],
                'layout': go.Layout(
                    title='Faculty Student Throughput - Undergraduate',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'}
                )
            }
        ),
        dcc.Graph(
            id='success-rate-bar-graph-sheet6',
            figure={
                'data': [
                    go.Bar(
                        x=[year for year in range(2014, 2023)],
                        y=df6.loc[df6['Department'] == dept, [str(year) for year in range(2014, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df6['Department']
                ],
                'layout': go.Layout(
                    title='Department Success Rates by Year (Sheet6)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='difference-graph-sheet6',
            figure={
                'data': [
                    go.Bar(
                        x=df6['Department'],
                        y=df6['Difference: 2014 vs 2022'],
                        name='Difference 2014 vs 2022',
                        marker=dict(color='blue')
                    )
                ],
                'layout': go.Layout(
                    title='Difference in Success Rates 2014 vs 2022 (Sheet6)',
                    xaxis={'title': 'Department'},
                    yaxis={'title': 'Difference (%)'}
                )
            }
        ),
        dcc.Graph(
            id='postgraduate-throughput-masters',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2015, 2023)],
                        y=df7.loc[df7['Department'] == dept, [str(year) for year in range(2015, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df7[df7['Department'].str.contains('Masters')]['Department']
                ],
                'layout': go.Layout(
                    title='Postgraduate Throughput - Masters',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='postgraduate-throughput-phd',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2015, 2023)],
                        y=df7.loc[df7['Department'] == dept, [str(year) for year in range(2015, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df7[df7['Department'].str.contains('PhD')]['Department']
                ],
                'layout': go.Layout(
                    title='Postgraduate Throughput - PhD',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Success Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='student-dropout-rates-undergraduate',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2014, 2023)],
                        y=df8.loc[df8['Department'] == dept, [str(year) for year in range(2014, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df8['Department']
                ],
                'layout': go.Layout(
                    title='Student Dropout Rates - Undergraduate',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Dropout Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='dropout-rate-first-year',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2014, 2021)],
                        y=df9.loc[df9['Department'] == dept, [str(year) for year in range(2014, 2021)]].values.flatten(),
                        name=dept
                    ) for dept in df9['Department']
                ],
                'layout': go.Layout(
                    title='Dropout Rate in The First Year',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Dropout Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='dropout-throughput-still-progress',
            figure={
                'data': [
                    go.Bar(
                        x=df10['Department'],
                        y=df10['Dropout'],
                        name='Dropout'
                    ),
                    go.Bar(
                        x=df10['Department'],
                        y=df10['Throughput'],
                        name='Throughput'
                    ),
                    go.Bar(
                        x=df10['Department'],
                        y=df10['Still in Progress'],
                        name='Still in Progress'
                    )
                ],
                'layout': go.Layout(
                    title='Dropout, Throughput, and Still in Progress (Sheet10)',
                    xaxis={'title': 'Department'},
                    yaxis={'title': 'Percentage (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='postgraduate-dropout-masters',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2014, 2023) if str(year) in df11.columns],
                        y=df11.loc[df11['Department'] == dept, [str(year) for year in range(2014, 2023) if str(year) in df11.columns]].values.flatten(),
                        name=dept
                    ) for dept in df11[df11['Department'].str.contains('Masters')]['Department']
                ],
                'layout': go.Layout(
                    title='Postgraduate Dropout - Masters',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Dropout Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='postgraduate-dropout-phd',
            figure={
                'data': [
                    go.Bar(
                        x=[str(year) for year in range(2014, 2023) if str(year) in df11.columns],
                        y=df11.loc[df11['Department'] == dept, [str(year) for year in range(2014, 2023) if str(year) in df11.columns]].values.flatten(),
                        name=dept
                    ) for dept in df11[df11['Department'].str.contains('PhD')]['Department']
                ],
                'layout': go.Layout(
                    title='Postgraduate Dropout - PhD',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Dropout Rate (%)'},
                    barmode='group'
                )
            }
        ),
        dcc.Graph(
            id='fas-graduation-rates',
            figure={
                'data': [
                    go.Scatter(
                        x=df12_filtered['Year'],
                        y=df12_filtered['Faculty'],
                        mode='lines+markers',
                        name='Actual Graduation Rate',
                        line=dict(color='blue')
                    ),
                    go.Scatter(
                        x=future_df12['Year'],
                        y=future_df12['Faculty'],
                        mode='lines+markers',
                        name='Forecasted Graduation Rate',
                        line=dict(color='red', dash='dash')
                    )
                ],
                'layout': go.Layout(
                    title='FAS Graduation Rates',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Graduation Rate (%)'},
                    annotations=[
                        dict(
                            x=2021,
                            y=df12.loc[df12['Year'] == '2021', 'Faculty'].values[0],
                            xref='x',
                            yref='y',
                            text='Difference: 2014 vs 2021 = 13%',
                            showarrow=True,
                            arrowhead=7,
                            ax=0,
                            ay=-40
                        )
                    ]
                )
            }
        ),
        dcc.Graph(
            id='graduation-rates-by-programme',
            figure={
                'data': [
                    go.Bar(
                        x=[year for year in range(2014, 2023)],
                        y=df13.loc[df13['Department'] == dept, [str(year) for year in range(2014, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df13['Department']
                ],
                'layout': go.Layout(
                    title='Graduation Rates By Programme',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Graduation Rate (%)'},
                    barmode='group',
                    annotations=[
                        dict(
                            x=2021,
                            y=df13[[str(year) for year in range(2014, 2023)]].max().max(),
                            xref='x',
                            yref='y',
                            text='Year with Most Graduates',
                            showarrow=True,
                            arrowhead=7,
                            ax=0,
                            ay=-40
                        )
                    ]
                )
            }
        ),
        dcc.Graph(
            id='difference-graph-sheet13',
            figure={
                'data': [
                    go.Bar(
                        x=df13['Department'],
                        y=df13['Difference: 2014 vs 2022'],
                        name='Difference 2014 vs 2022',
                        marker=dict(color='blue')
                    )
                ],
                'layout': go.Layout(
                    title='Difference in Graduation Rates 2014 vs 2022 (Sheet13)',
                    xaxis={'title': 'Department'},
                    yaxis={'title': 'Difference (%)'}
                )
            }
        ),
        dcc.Graph(
            id='postgraduate-graduation-rate',
            figure={
                'data': [
                    go.Scatter(
                        x=df14['Year'],
                        y=df14['Faculty'],
                        mode='lines+markers',
                        name='Actual Graduation Rate',
                        line=dict(color='blue')
                    ),
                    go.Scatter(
                        x=future_df14['Year'],
                        y=future_df14['Faculty'],
                        mode='lines+markers',
                        name='Forecasted Graduation Rate',
                        line=dict(color='red', dash='dash')
                    )
                ],
                'layout': go.Layout(
                    title='Postgraduate Graduation Rate',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Graduation Rate (%)'}
                )
            }
        ),
        dcc.Graph(
            id='pass-rates-department',
            figure={
                'data': [
                    go.Bar(
                        x=[year for year in range(2014, 2023)],
                        y=df15.loc[df15['Department'] == dept, [str(year) for year in range(2014, 2023)]].values.flatten(),
                        name=dept
                    ) for dept in df15['Department']
                ],
                'layout': go.Layout(
                    title='Pass Rates by Department (Sheet15)',
                    xaxis={'title': 'Year'},
                    yaxis={'title': 'Pass Rate (%)'},
                    barmode='group'
                )
            }
        )
    ]

    return graphs

# Staff page layout
bar_figures, line_figures, forecast_figures, fig_diff, fig_sheet5 = create_staff_charts()
staff_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Staff Preliminary Analysis"),
    
    html.Img(src='/assets/my_image.png', style={'width': '20%', 'height': 'auto'}),
    
    html.Div([
        html.H2("Faculty Data"),
        html.P("Reporting Period 2014 till 2022", style={'color': 'purple'})
    ], style={'textAlign': 'center'}),
    
    # Display the bar charts and forecast charts side by side
    html.Div([
        dcc.Graph(figure=bar_figures.get('Sheet1'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=bar_figures.get('Sheet2'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=bar_figures.get('Sheet3'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=forecast_figures.get('Sheet1'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=forecast_figures.get('Sheet2'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=forecast_figures.get('Sheet3'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=line_figures.get('Sheet4'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=fig_diff, style={'width': '50%', 'height': '600px'})  # Ensure Sheet5 chart is included
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=fig_sheet5, style={'width': '100%', 'height': '600px'})  # Full-width for Sheet5
    ], style={'display': 'flex', 'justifyContent': 'center'})
])

# Students page layout
fig1, fig2, fig3, fig4, fig5, fig6, fig7, fig8, fig9, fig10, fig11, fig12, fig13 = create_students_charts()
students_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Students Preliminary Analysis"),

    html.Div([
        dcc.Graph(figure=fig1)  # First chart
    ], style={'margin-bottom': '20px'}),  # Add space below the first chart

    html.Div([
        dcc.Graph(figure=fig2)  # Second chart
    ], style={'margin-bottom': '20px'}),  # Add space below the second chart

    html.Div([
        dcc.Graph(figure=fig3)  # Bar chart from Sheet2
    ], style={'margin-bottom': '20px'}),  # Add space below the third chart

    html.Div([
        dcc.Graph(figure=fig4)  # Bar chart from Sheet3
    ], style={'margin-bottom': '20px'}),  # Add space below the fourth chart

    html.Div([
        dcc.Graph(figure=fig5)  # Line chart from Sheet4
    ], style={'margin-bottom': '20px'}),  # Add space below the fifth chart

    html.Div([
        dcc.Graph(figure=fig6)  # Line chart from Sheet5
    ], style={'margin-bottom': '20px'}),  # Add space below the sixth chart

    html.Div([
        dcc.Graph(figure=fig7)  # Line chart from Sheet6
    ], style={'margin-bottom': '20px'}),  # Add space below the seventh chart

    html.Div([
        dcc.Graph(figure=fig8)  # Bar chart from Sheet7
    ], style={'margin-bottom': '20px'}),  # Add space below the eighth chart

    html.Div([
        dcc.Graph(figure=fig9)  # Line chart from Sheet8
    ], style={'margin-bottom': '20px'}),  # Add space below the ninth chart

    html.Div([
        dcc.Graph(figure=fig10)  # Bar chart for Difference 2014 vs 2022
    ], style={'margin-bottom': '20px'}),  # Add space below the tenth chart

    html.Div([
        dcc.Graph(figure=fig11)  # Bar chart from Sheet9
    ], style={'margin-bottom': '20px'}),  # Add space below the eleventh chart

    html.Div([
        dcc.Graph(figure=fig12)  # Bar chart from Sheet10
    ], style={'margin-bottom': '20px'}),  # Add space below the twelfth chart

    html.Div([
        dcc.Graph(figure=fig13)  # Bar chart from Sheet11
    ])
])

# Student Performance Indicators page layout
student_performance_graphs = create_student_performance_charts()
performance_layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("Student Performance Indicators"),
    *student_performance_graphs
])

# Define the main layout with a navigation bar and content
app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    html.Div([
        dcc.Link('Staff Preliminary Analysis', href='/'),
        html.Span(' | '),
        dcc.Link('Students Preliminary Analysis', href='/students'),
        html.Span(' | '),
        dcc.Link('Student Performance Indicators', href='/performance')
    ], style={'textAlign': 'center', 'margin': '20px'}),
    html.Div(id='page-content')
])

# Callback to render the appropriate page content
@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/students':
        return students_layout
    elif pathname == '/performance':
        return performance_layout
    else:
        return staff_layout

# Run the Dash app
if __name__ == '__main__':
    app.run_server(debug=True, port=8050)
